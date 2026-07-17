from doctr.models import ocr_predictor #Modelo utilizado para el reconocimiento de texto
from doctr.io import DocumentFile #utilizado para la lectura del documento a reconocer
import re # metodos para la busqueda, sustitucion y busqueda total(search,sub,findall)
from PIL import Image #recortes a una imagen
from openpyxl import load_workbook #leer y editar excel
from openpyxl.styles import PatternFill
from datetime import datetime #formato y leer fechas
import numpy as np #Conversion de imagen a matriz
import fitz #PyMuPDF Conversion de PDF a JPG
from IPython.display import display # Desplieque de imagen
import cv2 # Aplicar pre-procesado a la imagen

nombrePDF="Julio2026" #Nombre del PDF
expPeriodo="DEL 1 AL [1-31]"
expAnio="20[26-99]"
expMes="ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE"

# Abrir PDF
doc = fitz.open(f"PDF/{nombrePDF}.pdf") # Lectura del PDF
page = doc[0] # Seleccion de la hoja del archivo PDF
zoom = 2  # Renderizar como imagen (controlas resolución aquí)
mat = fitz.Matrix(zoom, zoom)
pix = page.get_pixmap(matrix=mat,dpi=600)
# Convertir a una imagen modelo RGB (3 Colores)
img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
#PREPROCESADO
img_np=np.asarray(img) #Conversion de la imagen a un arreglo bidimencional 3D
gris=cv2.cvtColor(img_np,cv2.COLOR_BGR2GRAY) #Aplicar filtro de gris
ruido=cv2.GaussianBlur(gris,(9,9),0) #Reduccion de ruido
thresh = cv2.threshold(ruido, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1] #Conversion a imagen binaria
thresh=np.asarray(thresh) #Conversion de la imagen en una matriz 2D
rgb_gray=cv2.cvtColor(thresh,cv2.COLOR_GRAY2RGB) # Conversion de una matriz 2D a una 3D
img=Image.fromarray(rgb_gray) #Conversion a imagen
# Coordenadas en imagen (ejemplo)
coordenadas={'datos':[100, 0, 5000,1100],'horas':[3130, 1200, 4820, 5700],'fechas':[340, 1200, 1220, 5700],'fecha_nota':[340, 1200, 2090, 5700]}

img_delimitada=img.crop(coordenadas['datos']) #Recorte de la imagen a la seccion Nombre
cabecera=np.asarray(img_delimitada)           #Conversion a matriz para el reconocimiento
#display(img_delimitada)                      #Desplegar imagen

img_delimitada=img.crop(coordenadas['horas'])
horas=np.asarray(img_delimitada)
#display(img_delimitada)

img_delimitada=img.crop(coordenadas['fechas'])
fechas=np.asarray(img_delimitada)
#display(img_delimitada)

img_delimitada=img.crop(coordenadas['fecha_nota'])
fecha_nota=np.asarray(img_delimitada)
#display(img_delimitada)

#cantidad de dias del mes. Aproximadamente 30s
modelo=ocr_predictor(det_arch='db_resnet50',reco_arch='crnn_vgg16_bn',pretrained=True,assume_straight_pages=True)#Modelo para la deteccion y reconocimiento de texto
result=modelo([fechas]) #Resultado
json_fechas=result.export() #Exportacion del resultado en formato json
#Identifica cada fecha del mes en el arreglo 'vector' y saca la cantidad de dias que tiene el mes
dias=[] #Arreglo donde se almacena cada dia ['miércoles, abril 01, 2026',...]
for pagina in json_fechas['pages']:
  for bloque in pagina['blocks']:
    for linea in bloque['lines']:
      linea=" ".join([w['value'] for w in linea['words']])
      linea=re.sub("â","a",linea)
      dias.append(linea)

if(dias!=31 or dias!=30 or dias!= 30): #Si la cantidad de dias que reconoce es diferente a 31,30 o 28 revisa campo a campo y elimina lo invalido
  for x in range(len(dias)-1):
    if(dias[x]=="."):
      dias.pop(x)

for x in range(len(dias)): #Agrega acentos a los dias
  dia=re.sub("sabado","sábado",dias[x])
  dias[x]=dia
  dia=re.sub("miercoles","miércoles",dias[x])
  dias[x]=dia

print(dias)

#Reconocimiento de todas las horas Entrada-Salida
result=modelo([horas])
json_horas=result.export()

horas=[] #Arreglo donde se guardaran las horas
for pagina in json_horas['pages']:
  for bloque in pagina['blocks']:
    for linea in bloque['lines']:
      linea="".join([w['value'] for w in linea['words']])
      horas.append(linea)
print(len(horas)/2)
print(horas)

#Funcion para reconocer Nombre, Mes, Año y Periodo
def reconocimiento():
  periodo=''
  # Coordenadas en imagen (ejemplo)
  result=modelo([cabecera]) #Resultado
  json_fechas=result.export() #Exportacion del resultado en formato json
  texto=[]
  for pagina in json_fechas['pages']:
    for bloque in pagina['blocks']:
      for linea in bloque['lines']:
        linea=" ".join([w['value'] for w in linea['words']])
        texto.append(linea)
  #'INMUEBLE:', 'SEDE TIJUANA', 'TECNICO:', 'Alex Rivera Perez', 'ANIO:', '2026', 'MES: MARZO', 'PERIODO:', 'DEL 1 AL31']
  for x in range(len(texto)):
    if(re.search("INMUEBLE:",texto[x])):
      inmueble=re.split(":",texto[x])
      if(inmueble[1]!=''):
        inmueble=inmueble[1]
        inmueble=re.sub(" ","",inmueble)
      else:
        inmueble=texto[x+1]
    if(re.search("TECNICO:",texto[x])):
      tecnico=re.split(":",texto[x])
      if(tecnico[1]!=''):
        tecnico=tecnico[1]
      else:
        tecnico=texto[x+1]
    if(re.search("ANO:",texto[x])): #'ANO:', '2026'
      anio=re.split(":",texto[x])
      #print(anio)
      if(anio[1]!=''):
        anio=anio[1]
      else:
        anio=texto[x+1]
    if(re.search("MES:",texto[x])):
      mes=re.split(":",texto[x])
      if(mes[1]!=''):
        mes=mes[1]
        mes=re.sub(" ","",mes)
      else:
        mes=texto[x+1]
    if(re.search("PERIODO:",texto[x])):
      periodo=re.split(":",texto[x]) #['PERIODO', '']
      if(periodo[1]!=''):
        periodo=periodo[1]
      else:
        periodo=texto[x+1]
    #Validaciones de Periodo
    if(periodo=='DEL1 AI31'):
      periodo='DEL 1 AL 31'
    if(periodo=='DEL1AL 31'):
      periodo='DEL 1 AL 31'
    if(periodo=='DEL 1 AL31'):
      periodo='DEL 1 AL 31'
    if(periodo=='DEL1AL30'):
      periodo='DEL 1 AL 30'
    if(periodo=='DEL1 AL 31'):
      periodo='DEL 1 AL 31'
  datos={
      'INMUEBLE':inmueble,
      'TECNICO':tecnico,
      'ANIO':anio,
      'MES':mes,
      'PERIODO':periodo
  }
  return datos

datos=reconocimiento()
nombre=datos['TECNICO']
anio=datos['ANIO']
mes=datos['MES']
periodo=datos['PERIODO']

dia_mes=[] #Arreglo para almacenar la fecha y la nota
result=modelo([fecha_nota]) # Reconocimiento de todas las fechas y notas
json_dia=result.export()

#Almacena la fecha y la nota si corresponde en una pocision diferente(['sabado,abril04,2026', 'NOLABORABLE'])
for pagina in json_dia['pages']:
  for bloque in pagina['blocks']:
    for linea in bloque['lines']:
      linea="".join([w['value'] for w in linea['words']])
      linea=re.sub("é","e",linea)
      linea=re.sub("â","a",linea)
      dia_mes.append(linea)
      #lo que se almacena en 'dia_mes' es la fecha y la nota en una posicion diferente(['domingo,febrero01,2026','NOLABORABLE'])

print(dia_mes)
print(len(dia_mes))

#Une el dia con la nota si corresponde (['sabado,abril04,2026,NOLABORABLE'])
eliminar=[]
c=0
for x in range (len(dia_mes)): #Recorre todos los elementos de [dia_mes]
#Si encuentra un NOLABORABLE en una posicion, lo une con la fecha anterior y deja en blanco la posicion
  if(re.search("NOLABORABLE",dia_mes[x])):
    dia_mes[x]=dia_mes[x-1]+","+dia_mes[x]
    c+=1
    dia_mes[x-1]=""
#Hace un recorrido por todo el arreglo, si encuentra una posicion vacia lo elimina
for y in range(len(dia_mes)):
  if(y>=len(dia_mes)):
    a=True
  else:
    if(dia_mes[y]==""):
      dia_mes.pop(y)
print(dia_mes)
print(len(dia_mes))

#Si hay notas escritas a mano, y si estan en pocisiones diferentes (['No','laborable']) las une en una sola pocision (['No laborable'])
fecha="lunes|martes|miercoles|jueves|viernes|sabado|domingo"
#print("total:",len(dia_mes))
for x in range(len(dia_mes)):
  for y in range(len(dia_mes)):
    if(y>=len(dia_mes)):
      a=True
    else:
      if(re.search(fecha,dia_mes[y])):
        a=True
      else:
        dia_mes[y-1]=dia_mes[y-1]+dia_mes[y]
        dia_mes.pop(y)
print(dia_mes)

#Formato de hora (11:00)
formatoHora="^[0-5][0-9]:[0-5][0-9]$"
corregir=[]

#Corregir hora y guardalo en arreglo 'horas' tanto entrada como salida
for n in range(len(horas)):
    depurado=horas[n]
    depurado=re.sub("\\.",":",depurado)
    depurado=re.sub(",",":",depurado)
    depurado=re.sub(";",":",depurado)
    depurado=re.sub("O","0",depurado)
    depurado=re.sub("o","0",depurado)
    depurado=re.sub("D","0",depurado)
    depurado=re.sub("\\(","1",depurado)
    depurado=re.sub("L","2",depurado)
    depurado=re.sub("Z","2",depurado)
    depurado=re.sub("I","1",depurado)
    depurado=re.sub("-","",depurado)
    depurado=re.sub("e","",depurado)
    depurado=re.sub("S","5",depurado)
    depurado=re.sub("s","5",depurado)
    depurado=re.sub("!","1",depurado)
    depurado=re.sub("C","0",depurado)
    depurado=re.sub("G","5",depurado)
    depurado=re.sub("i","1",depurado)
    corregir.append(horas[n])
    if(horas[n-1][0]=="2"):
        depurado=re.sub("l","1",depurado)
    else:
        depurado=re.sub("l","2",depurado)
    horas[n]=depurado
    #corregir -> 1:00
    if(re.search(formatoHora,depurado)):
      hora=re.split(":",depurado)
      if(re.search("^[1-5][0-9]:[1-5][0-9]$",hora[0])):
        horas[n]=depurado
      if(hora[0][0]=="1" and hora[0][1]=="6"):
        horas[n]=depurado
        nuevo=re.sub("6","0",hora[0])
        nuevo=nuevo+":"+hora[1]
        horas[n]=nuevo
        print(horas[n])
    else:
        nuevo=re.sub("a","",horas[n])
        horas[n]=nuevo
        if(re.search(":",horas[n])):
            partes=re.split(":",horas[n])
            if(partes[1][0]=="9"):
                nuevo=re.sub("9","4",partes[1])
                nuevo=partes[0]+":"+nuevo
                horas[n]=nuevo
            if(len(partes[0])==3):
              if(partes[0][1]==partes[0][2]):
                nuevo=partes[0][0]+partes[0][1]+":"+partes[1][0]+partes[1][1]
                horas[n]=nuevo
            if(len(partes[0])==1):
              if(partes[0]=="1"):
                nuevo="11:"+partes[1]
                horas[n]=nuevo
            if(len(partes[0])==2 and len(partes[1])==1):
              #EJEMPLO: 21:0
              if(re.search("[0-9]",partes[1]) and re.search("[0-9]{2}",partes[0])):
                nuevo=partes[0]+":"+partes[1][0]+partes[1][0]
                horas[n]=nuevo
        if(len(horas[n])==5): #SON 5 CARACTERES
          #Si tiene 5 caracteres con formato (hh'mm)
          if(re.search("'",horas[n][2])):
            nuevo=horas[n][0]+horas[n][1]+":"+horas[n][3]+horas[n][4]
            horas[n]=nuevo
          #Si tiene 5 caracteres con formato (hh3mm)
          if(re.search("3",horas[n][2])):
            nuevo=horas[n][0]+horas[n][1]+":"+horas[n][3]+horas[n][4]
            horas[n]=nuevo
          if(re.search("°",horas[n][2])):#hh°mm
            nuevo=horas[n][0]+horas[n][1]+":"+horas[n][3]+horas[n][4]
            horas[n]=nuevo
          #Si tiene 5 caracteres y el formato (hh:mm)
          if(re.search(":",horas[n])):
              hora=re.split(":",horas[n]) #Si hora tiene 5 caracteres (hh:mm) lo divide en 2 separado por :(hh y mm)
              if(re.search("[1-5][0-9]",hora[0])):
                if(hora[1][0]=="6"):
                  minutos=re.sub("6","0",hora[1])
                  horas[n]=hora[0]+":"+minutos
          if(horas[n][2]=="1"):
            #Si el 3er caracter es 1 (hh1hh) se reemplaza por : (hh:mm)
            nuevo=horas[n][0]+horas[n][1]+":"+horas[n][3]+horas[n][4]
            horas[n]=nuevo
          if(horas[n][2]=="0"):
            #Si el 3er caracter es 0 (hh0hh) se reemplaza por : (hh:mm)
            nuevo=horas[n][0]+horas[n][1]+":"+horas[n][3]+horas[n][4]
            horas[n]=nuevo
          if(horas[n][2]=="2"):
            #Si el 3er caracter es 2 (hh2hh) se reemplaza por : (hh:mm)
            nuevo=horas[n][0]+horas[n][1]+":"+horas[n][3]+horas[n][4]
            horas[n]=nuevo
          if(horas[n][2]=="5"):
            nuevo=horas[n][0]+horas[n][1]+":"+horas[n][3]+horas[n][4]
            horas[n]=nuevo
          if(horas[n][2]=="%"):
            nuevo=horas[n][0]+horas[n][1]+":"+horas[n][3]+horas[n][4]
            horas[n]=nuevo
        if(len(horas[n])==6):
            if(re.search("[0-9]{5}",horas[n])):
                if(horas[n][2]=="2"):
                    nuevo=horas[n][0]+horas[n][1]+":"+horas[n][3]+horas[n][4]
                    horas[n]=nuevo
        if(len(horas[n])==4):
            if(re.search("[0-9]{4}",horas[n])):
                nuevo=horas[n][0]+horas[n][1]+":"+horas[n][2]+horas[n][3]
                horas[n]=nuevo
        if(len(horas[n])==3):
            if(horas[n][0]=="3"):
                nuevo=re.sub("3","8",horas[n])
                horas[n]=nuevo
            if(horas[n][0]=="8"):
                nuevo=horas[n][0]+":"+horas[n][1]+horas[n][2]
                horas[n]=nuevo
print(horas)

#Llenar diccionario horario con fecha,entrada,salida y nota
horario={}
c=0
for x in range(len(dia_mes)):
  if(re.search("NO|No|no",dia_mes[x])==None):#Es dia laborable
    #print(x," a")
    if(c>=len(horas)): #Si NO hay elementos en horas[]
      #print(x)
      horario[x]={
        'Fecha':dias[x],
        'Nota':'',
        'Entrada':'',
        'Salida':''
      }
    else: #Si hay elementos en horas[]
      if(len(horas)%2!=0):# Si elementos en horas[] es numero impar
        #print(x," c")
        if(c<len(horas)-1): # Si aun hay elementos en horas[]
          #print(x, " ",c)
          horario[x]={
            'Fecha':dias[x],
            'Nota':'',
            'Entrada':horas[c],
            'Salida':horas[c+1]
          }
        else:
          horario[x]={
              'Fecha':dias[x],
              'Nota':'',
              'Entrada':horas[c],
              'Salida':''
          }
      else: #Si elementos en horas[] es numero par
        #print(x)
        horario[x]={
          'Fecha':dias[x],
          'Nota':'',
          'Entrada':horas[c],
          'Salida':horas[c+1]
        }
    c+=2
  else:
    #print(x," b")
    horario[x]={
      'Fecha':dias[x],
      'Nota':'NO LABORABLE',
      'Entrada':'',
      'Salida':''
    }
print(horario)

#from os import SCHED_RESET_ON_FORK
import os

horaValida="^(8|19|10|21):[0-5][0-9]$"
file=load_workbook(filename="Excel/formato.xlsx", data_only=True)
sheet=file['ASISTENCIA-Alex']
ren=12
#n_fecha=fecha.strftime("%d") #Extraer solo el dia (%d)
sheet.cell(column=6,row=7, value=nombre) #Nombre
sheet.cell(column=5,row=9, value=mes) #Mes
#Llenar excel con los datos del diccionario 'horario'
#num=12
color= PatternFill(start_color="F5C527", end_color="F5C527", fill_type="solid")
#sheet[f"A{num}"].fill=color
for x in range(len(dia_mes)):
  if(horario.get(x)):
    sheet.cell(column=1, row=ren, value=horario[x]['Fecha'])#Fecha
    sheet.cell(column=2, row=ren, value=horario[x]['Nota'])#Nota
    #Entrada
    #Hora de entrada es valida con el formato hh:mm
    if(re.search(horaValida,horario[x]['Entrada'])):
      #print("a")re.search(horaValida,horario[x]['Entrada'])
      sheet.cell(column=7, row=ren, value=horario[x]['Entrada'])#Hora de entrada
    else:
      #El formato de la hr no es valido
      if(horario[x]['Entrada']!=""):
        #Si no es valor vacio se subraya
        sheet[f"G{ren}"].fill=color
        sheet[f"G{ren}"].fill=color
        sheet.cell(column=7, row=ren, value=horario[x]['Entrada'])#Hora de entrada
    #Hora de salida es valida con formato hh:mm
    if(re.search(horaValida,horario[x]['Salida'])):
      sheet.cell(column=9, row=ren, value=horario[x]['Salida'])#Hora de salida
    else:
      #Formato de la hr no es valido y no tiene valor vacio
      if(horario[x]['Salida']!=""):
        #print("d")
        sheet[f"I{ren}"].fill=color
        sheet[f"I{ren}"].fill=color
        sheet.cell(column=9, row=ren, value=horario[x]['Salida'])#Hora de salida
    ren+=1
  elif(horario.get(x)==None):
    ren+=1
#Validacion de nombre
if(sheet.cell(column=6,row=7).value!="Alex Rivera Perez"):
  sheet[f"F7"].fill=color
#Validacion de periodo
if(re.search(expPeriodo,periodo)):
  sheet.cell(column=8,row=9, value=periodo) #Periodo
else:
  sheet.cell(column=8,row=9, value=periodo) #Periodo
  sheet[f"H9"].fill=color
#Validacion Año
if(re.search(expAnio,anio)):
  sheet.cell(column=2,row=9, value=anio) #Año
else:
  sheet.cell(column=2,row=9,value=anio) #Ano
  sheet[f"B9"].fill=color
#Validacion Mes
if(re.search(expMes,mes)):
  sheet.cell(column=5,row=9, value=mes) #Mes
else:
  sheet.cell(column=5,row=9,value=mes) #Mes
  sheet[f"E9"].fill=color
file=file.save(filename=f"Excel/ListaAsistencia,{nombrePDF} Alex.xlsx")
print("LISTO")

